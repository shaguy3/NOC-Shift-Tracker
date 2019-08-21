import openpyxl
from datetime import date, datetime, timedelta

HOURLY_RATE = 30
SHIFT_RATES = [1, 1.5, 1.5, 2, 2.5, 1.25, 1.5]
# [day, night, friday morning, saturday night, sabath, first overtime, second overtime]
DRIVES_PAY = 26
HEALTH_PAY = 1.2


def add_to_db(sheet):
    db_wb = openpyxl.load_workbook('db.xlsx')
    db_sheet = db_wb[db_wb.sheetnames[0]]
    for row in range(4, sheet.max_row + 1):
        if not sheet['E' + str(row)].value is None:
            db_sheet.append((sheet['A' + str(row)].value[-5:] + '/' + str(date.today().year)[2:],
                             sheet['E' + str(row)].value,
                             sheet['F' + str(row)].value))

    db_wb.save('db.xlsx')


def purge_db():
    db_wb = openpyxl.load_workbook('db.xlsx')
    db_sheet = db_wb[db_wb.sheetnames[0]]

    db_sheet.delete_cols(1, 3)

    db_sheet['A1'] = 'Date of shift'
    db_sheet['B1'] = 'Shift start'
    db_sheet['C1'] = 'Shift end'

    db_wb.save('db.xlsx')


# TODO: Write a function that will organize the db by date.
def organize_db():
    db_wb = openpyxl.load_workbook('db.xlsx')
    db_sheet = db_wb[db_wb.sheetnames[0]]

    for i in range(2, db_sheet.max_row):
        for j in (i, db_sheet.max_row):
            if int(db_sheet['A' + str(j)].value[6:]) < int(db_sheet['A' + str(j + 1)].value[6:]):
                temp_date = db_sheet['A' + str(j)].value
                db_sheet['A' + str(j)].value = db_sheet['A' + str(j + 1)].value
                db_sheet['A' + str(j + 1)].value = temp_date
            elif int(db_sheet['A' + str(j)].value[3:5]) < int(db_sheet['A' + str(j + 1)].value[3:5]):
                temp_date = db_sheet['A' + str(j)].value
                db_sheet['A' + str(j)].value = db_sheet['A' + str(j + 1)].value
                db_sheet['A' + str(j + 1)].value = temp_date
            elif int(db_sheet['A' + str(j)].value[:2]) < int(db_sheet['A' + str(j + 1)].value[:2]):
                temp_date = db_sheet['A' + str(j)].value
                db_sheet['A' + str(j)].value = db_sheet['A' + str(j + 1)].value
                db_sheet['A' + str(j + 1)].value = temp_date


def get_months_shifts(month):
    db_wb = openpyxl.load_workbook('db.xlsx')
    db_sheet = db_wb[db_wb.sheetnames[0]]
    this_months_shifts = []
    current_row = 2

    for row in range(current_row, db_sheet.max_row + 1):
        # print(f'Currently looking in row: {row}')
        # print(db_sheet['A' + str(current_row)].value[3:5])
        if db_sheet['A' + str(current_row)].value[3:5] == month:
            # print(f'Got shift at row: {row}')
            shift_date = db_sheet['A' + str(current_row)].value
            shift_start_time = db_sheet['B' + str(current_row)].value
            shift_end_time = db_sheet['C' + str(current_row)].value

            shift_start_datetime = datetime.strptime(shift_date + shift_start_time, '%d/%m/%y%H:%M')

            if int(shift_end_time[0:2]) < int(shift_start_time[0:2]):
                shift_end_datetime = datetime.strptime(shift_date, '%d/%m/%y') + \
                                     timedelta(days=1,
                                               hours=int(shift_end_time[0:2]),
                                               minutes=int(shift_end_time[3:]))
            else:
                shift_end_datetime = datetime.strptime(shift_date, '%d/%m/%y') + \
                                     timedelta(hours=int(shift_end_time[0:2]),
                                               minutes=int(shift_end_time[3:]))

            this_shift = [shift_start_datetime, shift_end_datetime]
            this_months_shifts.append(this_shift)
        current_row += 1

    return this_months_shifts


def split_to_minutes(shift):
    splitted_shift = []
    current_minute = shift[0] + timedelta(minutes=1)
    while current_minute <= shift[1]:
        splitted_shift.append(current_minute)
        current_minute += timedelta(minutes=1)

    return splitted_shift


def minute_cat(current_minute):
    if current_minute.weekday() == 4:
        if current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 6):
            return 2  # Night
        elif current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 16):
            return 3  # Friday morning
        else:
            return 5  # Saturday
    elif current_minute.weekday() == 5:
        if current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 17):
            return 5  # Saturday
        elif current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 22):
            return 4  # Saturday night
        else:
            return 2  # Night
    else:
        if current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 6):
            return 2  # Night
        elif current_minute < datetime(current_minute.year, current_minute.month, current_minute.day, 22):
            return 1  # Day
        else:
            return 2  # Night


def organize_shift(shift):
    splited_shift = split_to_minutes(shift)
    minutes_worked = 0  # For the overtime calculations
    shift_counters = [0, 0, 0, 0, 0, 0, 0]
    last_minute_cat = 0
    # [day, night, friday_morning, saturday_night, saturday, overtime, extended overtime]

    for current_minute in splited_shift:
        current_cat = minute_cat(current_minute)
        if minutes_worked >= 480:
            if minutes_worked >= 600:
                shift_counters[6] += 1
            elif last_minute_cat in [2, 3, 4, 5]:
                shift_counters[last_minute_cat - 1] += 1
            else:
                shift_counters[5] += 1
        else:
            shift_counters[current_cat - 1] += 1
            last_minute_cat = current_cat
        minutes_worked += 1

    return shift_counters


def month_calc(month_shifts, hourly_rate, shift_rates, drives_pay, health_pay):
    gross_pay = 0
    minute_rate = hourly_rate / 60
    for shift in month_shifts:
        organized_shift = organize_shift(shift)
        shift_pay = [a * b * minute_rate for a, b in zip(organized_shift, shift_rates)]
        gross_pay += sum(shift_pay)
        gross_pay += drives_pay
        gross_pay += health_pay
    print(gross_pay)


# TODO: Write a function that will calculate the net pay for a given month.


def main():

    # purge_db()

    # my_wb = openpyxl.load_workbook('may_2019.xlsx')
    # my_sheet = my_wb['Sheet1']
    # add_to_db(my_sheet)
    #
    may_shifts = get_months_shifts('05')
    # for may_shift in may_shifts:
    #     for minute in split_to_minutes(may_shift):
    #         print(minute, minute_cat(minute))

    month_calc(may_shifts, HOURLY_RATE, SHIFT_RATES, DRIVES_PAY, HEALTH_PAY)

    # purge_db()
    #
    # my_wb = openpyxl.load_workbook('to_add.xlsx')
    # my_sheet = my_wb[my_wb.sheetnames[0]]
    #
    # add_to_db(my_sheet)


if __name__ == '__main__':
    main()

# TODO: Write documentations.
















